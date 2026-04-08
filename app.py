import os
import pandas as pd
from flask import Flask, render_template, request, redirect, session, jsonify
from database import (conectar, obtenerusuarios, insertar_masivo,
                      buscar_estudiantes, editar_estudiante, eliminar_estudiante,
                      obtener_estudiantes_por_carrera)
from dashprincipal import creartablero
import database

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "clave_local_dev")

creartablero(app)


# ── HEADERS ANTI-CACHÉ ───────────────────────────────────────────────────────
@app.after_request
def add_header(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"]        = "no-cache"
    response.headers["Expires"]       = "0"
    return response


# ── LOGIN ────────────────────────────────────────────────────────────────────
@app.route("/", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        username = request.form.get("usuario",    "").strip()
        password = request.form.get("contraseña", "").strip()

        if not username or not password:
            error = "Por favor ingresa usuario y contraseña."
            return render_template("login.html", error=error)

        usuario = obtenerusuarios(username)

        if usuario and isinstance(usuario, dict):
            contrasena_bd = usuario.get("contraseña")
            rol           = usuario.get("rol", "estudiante").lower()

            if contrasena_bd and contrasena_bd == password:
                session["username"]        = usuario.get("nombre_usuario") or username
                session["rol"]             = rol
                session["nombre_completo"] = usuario.get("nombre_usuario", username)

                from database import registrar_log
                registrar_log(session["username"], "login", f"Inicio de sesión — rol: {rol}")

                if rol == "admin":       return redirect("/dashprincipal")
                elif rol == "profesor":  return redirect("/profesor")
                elif rol == "estudiante": return redirect("/estudiante")
                else:
                    error = "Rol no reconocido. Contacta al administrador."
            else:
                error = "Usuario o contraseña incorrectos."
        else:
            error = "Usuario o contraseña incorrectos."

    return render_template("login.html", error=error)


# ── HELPER ───────────────────────────────────────────────────────────────────
def redirect_seguro_por_rol():
    rol = session.get("rol")
    if rol == "admin":        return redirect("/dashprincipal")
    elif rol == "profesor":   return redirect("/profesor")
    elif rol == "estudiante": return redirect("/estudiante")
    else:
        session.clear()
        return redirect("/")


# ── ADMIN — dashboard principal ──────────────────────────────────────────────
@app.route("/dashprincipal")
def dashprincipal():
    if "username" not in session:
        return redirect("/")
    
    # Verificar que sea admin
    if session.get("rol") != "admin":
        return redirect_seguro_por_rol()
    
    # Para admin, obtener estadísticas globales (no filtradas por carrera)
    try:
        # Esta función ya existe en tu database.py
        data = database.obtener_stats_globales()
        mis_estadisticas = data["stats"]
        estadisticas_por_carrera = data["por_carrera"]
        
    except Exception as e:
        print(f"Error obteniendo estadísticas: {e}")
        mis_estadisticas = {
            "total_estudiantes": 0,
            "promedio_global": 0,
            "aprobados": 0,
            "reprobados": 0,
            "total_carreras": 0
        }
        estadisticas_por_carrera = []
    
    return render_template("dashprinci.html", 
                         usuario=session["username"],
                         stats=mis_estadisticas,
                         por_carrera=estadisticas_por_carrera)
# ── ADMIN — panel de gestión ──────────────────────────────────────────────────
@app.route("/admin")
def admin_panel():
    if "username" not in session:      return redirect("/")
    if session.get("rol") != "admin":  return redirect_seguro_por_rol()

    from database import obtener_stats_globales, obtener_todos_usuarios, \
                         obtener_log, obtener_alertas

    data = obtener_stats_globales()
    return render_template(
        "admin.html",
        usuario      = session["username"],
        stats        = data["stats"],
        por_carrera  = data["por_carrera"],
        usuarios_rol = data["usuarios_rol"],
        usuarios     = obtener_todos_usuarios(),
        log          = obtener_log(50),
        alertas      = obtener_alertas(),
    )


# ── ADMIN API — usuarios ──────────────────────────────────────────────────────
@app.route("/api/usuarios/listar")
def api_usuarios_listar():
    if "username" not in session or session.get("rol") != "admin":
        return jsonify({"ok": False, "error": "No autorizado"}), 403
    from database import obtener_todos_usuarios
    return jsonify({"ok": True, "usuarios": obtener_todos_usuarios()})


@app.route("/api/usuarios/crear", methods=["POST"])
def api_usuarios_crear():
    if "username" not in session or session.get("rol") != "admin":
        return jsonify({"ok": False, "error": "No autorizado"}), 403
    datos = request.get_json()
    try:
        from database import crear_usuario, registrar_log
        crear_usuario(datos)
        registrar_log(session["username"], "crear_usuario",
                      f"{datos['nombre_usuario']} — rol: {datos['rol']}")
        return jsonify({"ok": True})
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)})
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error: {str(e)}"})


@app.route("/api/usuarios/editar", methods=["POST"])
def api_usuarios_editar():
    if "username" not in session or session.get("rol") != "admin":
        return jsonify({"ok": False, "error": "No autorizado"}), 403
    datos = request.get_json()
    try:
        from database import editar_usuario, registrar_log
        editar_usuario(datos)
        registrar_log(session["username"], "editar_usuario",
                      f"ID:{datos['id_usuario']} — {datos['nombre_usuario']}")
        return jsonify({"ok": True})
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)})
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error: {str(e)}"})


@app.route("/api/usuarios/eliminar", methods=["POST"])
def api_usuarios_eliminar():
    if "username" not in session or session.get("rol") != "admin":
        return jsonify({"ok": False, "error": "No autorizado"}), 403
    datos = request.get_json()
    try:
        from database import eliminar_usuario, registrar_log
        eliminar_usuario(datos["id_usuario"])
        registrar_log(session["username"], "eliminar_usuario",
                      f"ID:{datos['id_usuario']}")
        return jsonify({"ok": True})
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)})
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error: {str(e)}"})


# ── ADMIN API — log ───────────────────────────────────────────────────────────
@app.route("/api/admin/log")
def api_log():
    if "username" not in session or session.get("rol") != "admin":
        return jsonify({"ok": False, "error": "No autorizado"}), 403
    from database import obtener_log
    return jsonify({"ok": True, "log": obtener_log(100)})


# ── ADMIN API — alertas ───────────────────────────────────────────────────────
@app.route("/api/admin/alertas")
def api_alertas():
    if "username" not in session or session.get("rol") != "admin":
        return jsonify({"ok": False, "error": "No autorizado"}), 403
    from database import obtener_alertas
    return jsonify({"ok": True, "alertas": obtener_alertas()})


# ── PROFESOR ─────────────────────────────────────────────────────────────────
@app.route("/profesor")
def profesor_dashboard():
    if "username" not in session:        return redirect("/")
    if session.get("rol") != "profesor": return redirect_seguro_por_rol()

    usuario = obtenerusuarios(session["username"])
    if not usuario:
        session.clear()
        return redirect("/")

    carrera_profesor = usuario.get("carrera")
    if not carrera_profesor:
        return "El profesor no tiene asignada ninguna carrera.", 400

    estudiantes_carrera = obtener_estudiantes_por_carrera(carrera_profesor)
    return render_template(
        "profesor.html",
        usuario    = session["username"],
        carrera    = carrera_profesor,
        estudiantes= estudiantes_carrera,
    )


@app.route("/profesor/exportar_excel")
def exportar_excel_profesor():
    if "username" not in session:          return redirect("/")
    if session.get("rol") != "profesor":   return redirect_seguro_por_rol()

    from flask import send_file
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    usuario          = obtenerusuarios(session["username"])
    carrera_profesor = usuario.get("carrera")
    if not carrera_profesor:
        return "Sin carrera asignada", 400

    estudiantes = obtener_estudiantes_por_carrera(carrera_profesor)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estudiantes"

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1A1A18")
    center       = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        left=Side(style="thin", color="E8E8E4"),
        right=Side(style="thin", color="E8E8E4"),
        top=Side(style="thin", color="E8E8E4"),
        bottom=Side(style="thin", color="E8E8E4"),
    )

    headers = ["Nombre", "Edad", "Carrera", "Nota 1", "Nota 2", "Nota 3", "Promedio", "Nivel"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = thin_border

    fills = {
        "excepcional": PatternFill("solid", fgColor="FDF5E0"),
        "bueno":       PatternFill("solid", fgColor="EAFAF1"),
        "regular":     PatternFill("solid", fgColor="FEFAEC"),
        "malo":        PatternFill("solid", fgColor="FDF0EE"),
    }

    for e in estudiantes:
        p = float(e.get("promedio") or 0)
        if p == 5.0:   nivel, fk = "Excepcional", "excepcional"
        elif p >= 4.0: nivel, fk = "Bueno",       "bueno"
        elif p >= 3.0: nivel, fk = "Regular",     "regular"
        else:          nivel, fk = "Malo",         "malo"

        ws.append([e.get("nombre",""), e.get("edad",""), e.get("carrera",""),
                   e.get("nota1",""), e.get("nota2",""), e.get("nota3",""), p, nivel])
        row = ws.max_row
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.alignment = center; cell.border = thin_border
            if col in (7, 8): cell.fill = fills[fk]

    for i, w in enumerate([28,8,20,9,9,9,12,14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22

    buffer = io.BytesIO()
    wb.save(buffer); buffer.seek(0)
    nombre_archivo = f"estudiantes_{carrera_profesor.replace(' ','_')}.xlsx"
    return send_file(buffer,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=nombre_archivo)


# ── ESTUDIANTE ───────────────────────────────────────────────────────────────
@app.route("/estudiante")
def estudiante_dashboard():
    if "username" not in session:          return redirect("/")
    if session.get("rol") != "estudiante": return redirect_seguro_por_rol()

    usuario = obtenerusuarios(session["username"])
    if not usuario:
        session.clear()
        return redirect("/")

    from database import obtener_historial, obtener_stats_carrera, obtener_posicion_carrera

    id_estudiante    = usuario.get("id_estudiante")
    carrera          = usuario.get("carrera_estudiante") or usuario.get("carrera")
    promedio         = float(usuario.get("promedio") or 0)
    historial        = obtener_historial(id_estudiante) if id_estudiante else []
    stats            = obtener_stats_carrera(carrera)   if carrera       else {}
    posicion         = obtener_posicion_carrera(id_estudiante, carrera) if id_estudiante and carrera else None
    promedio_carrera = float(stats.get("promedio_carrera") or 0) if stats else 0
    diferencia       = round(promedio - promedio_carrera, 2)

    return render_template(
        "student_dashboard.html",
        usuario    = session["username"],
        estudiante = {
            "nombre":   usuario.get("nombre_estudiante") or session["username"],
            "edad":     usuario.get("edad"),
            "carrera":  carrera,
            "nota1":    float(usuario.get("nota1") or 0),
            "nota2":    float(usuario.get("nota2") or 0),
            "nota3":    float(usuario.get("nota3") or 0),
            "promedio": promedio,
        },
        historial  = historial,
        stats      = stats,
        posicion   = posicion,
        diferencia = diferencia,
    )


@app.route("/estudiante/reporte_pdf")
def descargar_reporte():
    if "username" not in session:          return redirect("/")
    if session.get("rol") != "estudiante": return redirect_seguro_por_rol()

    from database import obtener_historial, obtener_stats_carrera, obtener_posicion_carrera
    from reporte_pdf import generar_pdf
    from flask import send_file

    usuario          = obtenerusuarios(session["username"])
    id_estudiante    = usuario.get("id_estudiante")
    carrera          = usuario.get("carrera_estudiante") or usuario.get("carrera")
    promedio         = float(usuario.get("promedio") or 0)
    historial        = obtener_historial(id_estudiante) if id_estudiante else []
    stats            = obtener_stats_carrera(carrera)   if carrera       else {}
    posicion         = obtener_posicion_carrera(id_estudiante, carrera) if id_estudiante and carrera else None
    promedio_carrera = float((stats or {}).get("promedio_carrera") or 0)
    diferencia       = round(promedio - promedio_carrera, 2)

    estudiante = {
        "nombre":   usuario.get("nombre_estudiante") or session["username"],
        "carrera":  carrera,
        "nota1":    float(usuario.get("nota1") or 0),
        "nota2":    float(usuario.get("nota2") or 0),
        "nota3":    float(usuario.get("nota3") or 0),
        "promedio": promedio,
    }

    pdf            = generar_pdf(estudiante, historial, stats, posicion, diferencia)
    nombre_archivo = f"reporte_{estudiante['nombre'].replace(' ','_')}.pdf"
    return send_file(pdf, mimetype="application/pdf",
                     as_attachment=True, download_name=nombre_archivo)


# ── TOP ESTUDIANTES ──────────────────────────────────────────────────────────
@app.route("/top_estudiantes")
def top_estudiantes():
    if "username" not in session: return redirect("/")
    return render_template("top_estudiantes.html")

@app.route("/api/estudiantes/top")
def api_top():
    if "username" not in session:
        return jsonify({"ok": False, "error": "No autenticado"}), 401
    try:
        from database import obtener_top_estudiantes
        estudiantes = obtener_top_estudiantes(10)
        for e in estudiantes:
            e['promedio'] = float(e['promedio'])
        return jsonify({"ok": True, "estudiantes": estudiantes})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# ── EDITAR ESTUDIANTE ────────────────────────────────────────────────────────
@app.route("/editar_estudiante")
def editar_estudiante_page():
    if "username" not in session: return redirect("/")
    return render_template("editar_estudiante.html")

@app.route("/api/estudiantes/buscar")
def api_buscar():
    if "username" not in session:
        return jsonify({"ok": False, "error": "No autenticado"}), 401
    nombre = request.args.get("nombre", "")
    try:
        return jsonify({"ok": True, "estudiantes": buscar_estudiantes(nombre)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "estudiantes": []})

@app.route("/api/estudiantes/editar", methods=["POST"])
def api_editar():
    if "username" not in session:
        return jsonify({"ok": False, "error": "No autenticado"}), 401
    datos = request.get_json()
    try:
        editar_estudiante(datos)
        from database import registrar_log
        registrar_log(session["username"], "editar_estudiante",
                      f"ID:{datos['id_estudiante']} — {datos['nombre']}")
        return jsonify({"ok": True})
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)})
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error al actualizar: {str(e)}"})

@app.route("/api/estudiantes/eliminar", methods=["POST"])
def api_eliminar():
    if "username" not in session:
        return jsonify({"ok": False, "error": "No autenticado"}), 401
    datos = request.get_json()
    try:
        eliminar_estudiante(datos["id_estudiante"])
        from database import registrar_log
        registrar_log(session["username"], "eliminar_estudiante",
                      f"ID:{datos['id_estudiante']}")
        return jsonify({"ok": True})
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)})
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error al eliminar: {str(e)}"})


# ── CARGA MASIVA ─────────────────────────────────────────────────────────────
@app.route("/carga_masiva", methods=["GET", "POST"])
def carga_masiva():
    if "username" not in session: return redirect("/")
    if request.method == "GET":
        return render_template("carga_masiva.html")

    if "archivo" not in request.files:
        return jsonify({"ok": False, "error": "No se recibió ningún archivo."})
    archivo = request.files["archivo"]
    if archivo.filename == "":
        return jsonify({"ok": False, "error": "No seleccionaste ningún archivo."})
    if not archivo.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"ok": False, "error": "El archivo debe ser .xlsx o .xls"})

    try:
        df        = pd.read_excel(archivo)
        resultado = insertar_masivo(df)
        from database import registrar_log
        registrar_log(session["username"], "carga_masiva",
                      f"Insertados:{resultado['insertados']} Duplicados:{resultado['duplicados']} Invalidos:{resultado['invalidos']}")
        return jsonify({
            "ok": True,
            "insertados": int(resultado["insertados"]),
            "duplicados": int(resultado["duplicados"]),
            "vacios":     int(resultado["vacios"]),
            "invalidos":  int(resultado["invalidos"]),
            "detalles":   resultado["errores"],
        })
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e), "detalles": []})
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error procesando el archivo: {str(e)}", "detalles": []})


# ── JUEGO ────────────────────────────────────────────────────────────────────
@app.route("/juego")
def juego():
    return render_template("gallina-pro.html")


# ── LOGOUT ───────────────────────────────────────────────────────────────────
@app.route("/logout")
def logout():
    from database import registrar_log
    if "username" in session:
        registrar_log(session["username"], "logout", "Cierre de sesión")
    session.clear()
    return redirect("/")


# ── 404 ──────────────────────────────────────────────────────────────────────
@app.errorhandler(404)
def pagina_no_encontrada(e):
    return render_template("404.html"), 404


if __name__ == "__main__":
    app.run(debug=True)